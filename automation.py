from selenium import webdriver 
from selenium.webdriver.common.by import By 
import openpyxl 

class GameDataCollector:

    def __init__(self):
        self.driver = webdriver.Chrome()
        self.workboot = openpyxl.workboot()

    def collect_data(self, url):
        self.driver.get(url)
        titles = self.driver.find_elements(By.XPATH, "//h3[@class='product-name']")
        prices = self.driver.find_elements(By.XPATH, "//span[@class='price-boleto']")
        return zip(titles, prices)

    def add_to_sheet(self, data, sheet):
        for title, price in data:
            sheet.append([title.text, price.text])

    def create_sheet(self, name):
        self.workboot.create_sheet(name)
        sheet = self.workboot[name]
        sheet['A1'].value = 'Game'
        sheet['B1'].value = 'Price'
        return sheet

    def save(self, filename):
        self.workboot.save(filename)

collector = GameDataCollector()

sheet_games_ps4 = collector.create_sheet('Games PS4')
dataPS4_1 = collector.collect_data('https://www.lojasgamemania.com.br/jogo-playstation-4')
collector.add_to_sheet(dataPS4_1, sheet_games_ps4)

dataPS4_2 = collector.collect_data('https://www.lojasgamemania.com.br/jogo-playstation-4?p=2')
collector.add_to_sheet(dataPS4_2, sheet_games_ps4)

sheet_games_sw = collector.create_sheet('Games SWITCH')
dataSW = collector.collect_data('https://www.lojasgamemania.com.br/jogos-de-nintendo-switch')
collector.add_to_sheet(dataSW, sheet_games_sw)

collector.save('Games.xlsx')
